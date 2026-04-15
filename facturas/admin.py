from django.contrib import admin
from django.contrib.auth.models import User
from django.http import HttpResponse
from django.urls import path
from django.template.response import TemplateResponse
from django.db.models import Sum, Count
from django.db.models.functions import TruncMonth, TruncYear
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from .models import Factura


# ─────────────────────────────────────────────
# ACCIÓN: Exportar a Excel (con columna usuario)
# ─────────────────────────────────────────────
def exportar_excel(modeladmin, request, queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Facturas"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1a73e8")
    total_fill = PatternFill("solid", fgColor="E8F0FE")
    center = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Encabezados — ahora incluye Usuario
    columnas = [
        ("Usuario", 20),
        ("Proveedor", 30),
        ("RUC", 15),
        ("Timbrado", 15),
        ("Fecha Emisión", 18),
        ("Importe Total (₲)", 20),
        ("Tipo", 22),
        ("Estado", 15),
        ("Notas", 40),
        ("Fecha Carga", 22),
    ]

    for col_num, (titulo, ancho) in enumerate(columnas, start=1):
        celda = ws.cell(row=1, column=col_num, value=titulo)
        celda.font = header_font
        celda.fill = header_fill
        celda.alignment = center
        celda.border = border
        ws.column_dimensions[celda.column_letter].width = ancho

    ws.row_dimensions[1].height = 25

    # Datos
    for row_num, f in enumerate(queryset, start=2):
        usuario_nombre = f.usuario.get_full_name() or f.usuario.username if f.usuario else "Sin usuario"
        valores = [
            usuario_nombre,
            f.nombre_proveedor,
            f.ruc,
            f.timbrado,
            str(f.fecha_emision) if f.fecha_emision else "",
            float(f.importe_total) if f.importe_total else 0,
            f.get_tipo_display(),
            f.get_estado_display(),
            f.notas,
            f.creado.strftime("%Y-%m-%d %H:%M") if f.creado else "",
        ]
        for col_num, valor in enumerate(valores, start=1):
            celda = ws.cell(row=row_num, column=col_num, value=valor)
            celda.border = border
            celda.alignment = Alignment(vertical="center")

    # Fila de total
    fila_total = queryset.count() + 2
    total = queryset.aggregate(Sum('importe_total'))['importe_total__sum'] or 0
    ws.cell(row=fila_total, column=5, value="TOTAL").font = Font(bold=True)
    celda_total = ws.cell(row=fila_total, column=6, value=float(total))
    celda_total.font = Font(bold=True)
    celda_total.fill = total_fill
    celda_total.border = border

    # Nombre de archivo inteligente
    usuarios = queryset.values_list('usuario__username', flat=True).distinct()
    if usuarios.count() == 1:
        sufijo = usuarios.first() or "sin_usuario"
    else:
        sufijo = "todos"
    nombre = f"facturas_{sufijo}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{nombre}"'
    wb.save(response)
    return response

exportar_excel.short_description = "📥 Exportar seleccionados a Excel"


# ─────────────────────────────────────────────
# ADMIN
# ─────────────────────────────────────────────
@admin.register(Factura)
class FacturaAdmin(admin.ModelAdmin):

    actions = [exportar_excel]

    list_display = [
        'get_usuario',
        'nombre_proveedor',
        'timbrado',
        'ruc',
        'fecha_emision',
        'importe_total',
        'tipo',
        'estado',
        'creado',
    ]

    list_filter = ['usuario', 'tipo', 'estado', 'fecha_emision']
    search_fields = ['nombre_proveedor', 'timbrado', 'ruc', 'notas', 'usuario__username']
    readonly_fields = ['texto_ocr', 'creado', 'actualizado']

    fieldsets = (
        ('Usuario', {'fields': ('usuario',)}),
        ('Imagen', {'fields': ('imagen',)}),
        ('Tipo de Factura', {'fields': ('tipo',)}),
        ('Datos de la Factura', {
            'fields': ('fecha_emision', 'timbrado', 'ruc', 'nombre_proveedor', 'importe_total', 'estado')
        }),
        ('Información Adicional', {'fields': ('notas', 'texto_ocr')}),
        ('Fechas del Sistema', {
            'fields': ('creado', 'actualizado'),
            'classes': ('collapse',)
        }),
    )

    # Columna usuario con link
    def get_usuario(self, obj):
        if obj.usuario:
            return obj.usuario.get_full_name() or obj.usuario.username
        return "—"
    get_usuario.short_description = "Usuario"
    get_usuario.admin_order_field = "usuario__username"

    # ── URLs personalizadas ──
    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path('dashboard/', self.admin_site.admin_view(self.dashboard_view), name='facturas_dashboard'),
            path('dashboard/excel/', self.admin_site.admin_view(self.exportar_excel_usuario), name='facturas_excel_usuario'),
        ]
        return custom + urls

    # ── Vista Dashboard con filtro por usuario ──
    def dashboard_view(self, request):
        MESES = ['Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic']
        anio_actual = timezone.now().year

        # Filtro por usuario
        usuario_id = request.GET.get('usuario', '')
        usuarios = User.objects.filter(facturas__isnull=False).distinct().order_by('username')
        qs = Factura.objects.all()
        usuario_seleccionado = None

        if usuario_id:
            try:
                usuario_seleccionado = User.objects.get(pk=usuario_id)
                qs = qs.filter(usuario=usuario_seleccionado)
            except User.DoesNotExist:
                pass

        # Resumen general
        total_facturas = qs.count()
        total_importe = qs.aggregate(Sum('importe_total'))['importe_total__sum'] or 0
        pendientes = qs.filter(estado='pendiente').count()
        pagadas = qs.filter(estado='pagada').count()
        vencidas = qs.filter(estado='vencida').count()

        # Resumen anual
        resumen_anual = (
            qs.filter(fecha_emision__isnull=False)
            .annotate(anio=TruncYear('fecha_emision'))
            .values('anio')
            .annotate(total=Sum('importe_total'), cantidad=Count('id'))
            .order_by('-anio')
        )

        # Resumen mensual
        resumen_mensual = (
            qs.filter(fecha_emision__isnull=False, fecha_emision__year=anio_actual)
            .annotate(mes=TruncMonth('fecha_emision'))
            .values('mes')
            .annotate(total=Sum('importe_total'), cantidad=Count('id'))
            .order_by('mes')
        )

        # Resumen por usuario (solo si no hay filtro)
        resumen_usuarios = []
        if not usuario_id:
            resumen_usuarios = (
                Factura.objects
                .values('usuario__username', 'usuario__first_name', 'usuario__last_name')
                .annotate(total=Sum('importe_total'), cantidad=Count('id'))
                .order_by('-total')
            )

        # Datos Chart.js
        meses_labels = []
        meses_valores = []
        for item in resumen_mensual:
            meses_labels.append(MESES[item['mes'].month - 1])
            meses_valores.append(float(item['total'] or 0))

        estado_labels = ['Pendiente', 'Pagada', 'Vencida']
        estado_valores = [pendientes, pagadas, vencidas]

        context = {
            **self.admin_site.each_context(request),
            'title': 'Dashboard de Facturas',
            'total_facturas': total_facturas,
            'total_importe': total_importe,
            'pendientes': pendientes,
            'pagadas': pagadas,
            'vencidas': vencidas,
            'resumen_anual': resumen_anual,
            'resumen_mensual': resumen_mensual,
            'resumen_usuarios': resumen_usuarios,
            'meses_labels': meses_labels,
            'meses_valores': meses_valores,
            'estado_labels': estado_labels,
            'estado_valores': estado_valores,
            'anio_actual': anio_actual,
            'MESES': MESES,
            'usuarios': usuarios,
            'usuario_id': usuario_id,
            'usuario_seleccionado': usuario_seleccionado,
        }

        return TemplateResponse(request, 'admin/facturas/dashboard.html', context)

    # ── Exportar Excel filtrado por usuario desde dashboard ──
    def exportar_excel_usuario(self, request):
        usuario_id = request.GET.get('usuario', '')
        qs = Factura.objects.all().order_by('-fecha_emision')
        sufijo = "todos"

        if usuario_id:
            try:
                usuario_obj = User.objects.get(pk=usuario_id)
                qs = qs.filter(usuario=usuario_obj)
                sufijo = usuario_obj.username
            except User.DoesNotExist:
                pass

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Facturas"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="1a73e8")
        total_fill = PatternFill("solid", fgColor="E8F0FE")
        center = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        columnas = [
            ("Usuario", 20),
            ("Proveedor", 30),
            ("RUC", 15),
            ("Timbrado", 15),
            ("Fecha Emisión", 18),
            ("Importe Total (₲)", 20),
            ("Tipo", 22),
            ("Estado", 15),
            ("Notas", 40),
            ("Fecha Carga", 22),
        ]

        for col_num, (titulo, ancho) in enumerate(columnas, start=1):
            celda = ws.cell(row=1, column=col_num, value=titulo)
            celda.font = header_font
            celda.fill = header_fill
            celda.alignment = center
            celda.border = border
            ws.column_dimensions[celda.column_letter].width = ancho

        ws.row_dimensions[1].height = 25

        for row_num, f in enumerate(qs, start=2):
            usuario_nombre = f.usuario.get_full_name() or f.usuario.username if f.usuario else "Sin usuario"
            valores = [
                usuario_nombre,
                f.nombre_proveedor,
                f.ruc,
                f.timbrado,
                str(f.fecha_emision) if f.fecha_emision else "",
                float(f.importe_total) if f.importe_total else 0,
                f.get_tipo_display(),
                f.get_estado_display(),
                f.notas,
                f.creado.strftime("%Y-%m-%d %H:%M") if f.creado else "",
            ]
            for col_num, valor in enumerate(valores, start=1):
                celda = ws.cell(row=row_num, column=col_num, value=valor)
                celda.border = border
                celda.alignment = Alignment(vertical="center")

        fila_total = qs.count() + 2
        total = qs.aggregate(Sum('importe_total'))['importe_total__sum'] or 0
        ws.cell(row=fila_total, column=5, value="TOTAL").font = Font(bold=True)
        celda_total = ws.cell(row=fila_total, column=6, value=float(total))
        celda_total.font = Font(bold=True)
        celda_total.fill = total_fill
        celda_total.border = border

        nombre = f"facturas_{sufijo}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{nombre}"'
        wb.save(response)
        return response

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context['dashboard_url'] = '/admin/facturas/factura/dashboard/'
        return super().changelist_view(request, extra_context=extra_context)