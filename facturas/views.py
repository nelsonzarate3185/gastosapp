# facturas/views.py

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.parsers import MultiPartParser, FormParser
from django.conf import settings
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.db.models import Sum, Q
from django.utils import timezone
from .models import Factura, Ingreso
from .serializers import FacturaSerializer, IngresoSerializer, UsuarioSerializer
import requests
import re
import csv
import io
from django.http import HttpResponse
from django.db.models.functions import TruncMonth


def _puede_ver_todo(user):
    """True si el usuario puede ver transacciones de todos los usuarios."""
    return user.is_superuser or user.has_perm('facturas.ver_todo')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ══════════════════════════════════════════════════════
# OCR - Sin cambios en la lógica, solo se mantiene
# ══════════════════════════════════════════════════════

def ocr_con_api(imagen):
    api_key = settings.OCR_API_KEY
    try:
        response = requests.post(
            'https://api.ocr.space/parse/image',
            files={'image': imagen},
            data={
                'apikey': api_key,
                'language': 'spa',
                'isOverlayRequired': False,
            },
            timeout=30,
        )
    except requests.exceptions.Timeout:
        raise Exception('El servicio OCR tardó demasiado. Intentá de nuevo.')
    except requests.exceptions.ConnectionError:
        raise Exception('No se pudo conectar al servicio OCR. Verificá la conexión.')

    try:
        result = response.json()
    except ValueError:
        raise Exception(
            f'El servicio OCR devolvió una respuesta inválida (HTTP {response.status_code}). '
            'Es posible que se haya superado el límite de uso del API key gratuito.'
        )

    if result.get('IsErroredOnProcessing'):
        raise Exception(result.get('ErrorMessage', 'Error en OCR'))
    parsed = result.get('ParsedResults', [])
    if not parsed:
        return ''
    return parsed[0].get('ParsedText', '')


def _limpiar_monto(s):
    """'1.901.004' → 1901004.0  |  '220.000' → 220000.0"""
    if not s:
        return None
    s = str(s).strip()
    # Si tiene más de un punto, son separadores de miles
    if s.count('.') > 1:
        s = s.replace('.', '').replace(',', '')
    else:
        s = s.replace('.', '').replace(',', '')
    try:
        return float(s)
    except ValueError:
        return None


def extraer_datos_ocr(texto):
    """Extracción regex de campos de facturas paraguayas."""
    from datetime import datetime as _dt

    datos = {
        'texto_ocr':        texto,
        'nombre_proveedor': '',
        'timbrado':         '',
        'ruc':              '',
        'numero_factura':   '',
        'importe_total':    None,
        'importe_impuesto': None,
        'fecha_emision':    None,
    }

    if not texto:
        return datos

    texto_upper = texto.upper()
    lineas_orig = [l.strip() for l in texto.split('\n') if l.strip()]

    # ── RUC del EMISOR ────────────────────────────────────────────
    # "RUC: 4934368 - 8"  |  "RUC: 80014137-7"
    ruc_m = re.search(r'RUC[^:\d\n]{0,10}[:\s]+(\d{5,8})\s*[-–]\s*(\d)', texto_upper)
    if not ruc_m:
        ruc_m = re.search(r'\b(\d{7,8})\s*[-–]\s*(\d)\b', texto)
    if ruc_m:
        datos['ruc'] = f"{ruc_m.group(1)}-{ruc_m.group(2)}"

    # ── Timbrado ──────────────────────────────────────────────────
    # "TIMBRADO N° 18733343"  |  "Timbrado Nº: 16087162"
    tim_m = re.search(r'TIMBRADO[^:\d\n]{0,6}[N°Nº#:\s]+(\d{7,10})', texto_upper)
    if tim_m:
        datos['timbrado'] = tim_m.group(1)
    else:
        # Número de 8 dígitos standalone (formato más común PY)
        tim2 = re.search(r'\b(\d{8})\b', texto)
        if tim2:
            datos['timbrado'] = tim2.group(1)

    # ── Número de factura ─────────────────────────────────────────
    # "001-001-0000508"  |  "001-001 0000508"  |  "016-004-0034087"
    nro_m = re.search(r'\b(\d{3})\s*[-]\s*(\d{3})\s*[-\s]\s*(\d{7})\b', texto)
    if nro_m:
        datos['numero_factura'] = f"{nro_m.group(1)}-{nro_m.group(2)}-{nro_m.group(3)}"

    # ── Nombre del proveedor ──────────────────────────────────────
    TIPOS = r'S\.A\.E\.C\.A\.|S\.R\.L\.|S\.A\.|LTDA\.?|E\.I\.R\.L\.|S\.C\.S\.|SAECA'
    for linea in lineas_orig[:15]:
        if re.search(TIPOS, linea.upper()):
            datos['nombre_proveedor'] = linea.strip()
            break

    if not datos['nombre_proveedor']:
        razon_m = re.search(r'(?:NOMBRE\s+O\s+)?RAZ[ÓO]N\s+SOCIAL[:\s]+(.+)', texto_upper)
        if razon_m:
            datos['nombre_proveedor'] = razon_m.group(1).strip().title()

    if not datos['nombre_proveedor']:
        # Primera línea significativa que no sea un encabezado conocido
        skip = re.compile(r'^(RUC|TIMBRADO|FACTURA|FECHA|KUDE|KuDE|INICIO|FIN\s+DE|N[°º]|TEL[EÉ]F|EMAIL|DIRECCI)', re.I)
        for linea in lineas_orig[:8]:
            if len(linea) > 6 and not skip.match(linea):
                datos['nombre_proveedor'] = linea.strip()
                break

    # ── Importe total ─────────────────────────────────────────────
    patrones_total = [
        r'TOTAL\s+(?:A\s+PAGAR\s+GUARAN[IÍ]ES?\b[^0-9]{0,40})([\d][.\d\s]+)',
        r'TOTAL\s+DE\s+LA\s+OPERACI[ÓO]N[^\d]{0,10}([\d][.\d]+)',
        r'TOTAL\s+A\s+PAGAR[^\d]{0,10}([\d][.\d]+)',
        r'TOTAL\s+PAGAR[^\d]{0,10}([\d][.\d]+)',
    ]
    for pat in patrones_total:
        m = re.search(pat, texto_upper)
        if m:
            val = _limpiar_monto(m.group(1).replace(' ', ''))
            if val:
                datos['importe_total'] = val
                break

    if datos['importe_total'] is None:
        # Fallback: número más grande en formato guaraní NNN.NNN o NNNN.NNN
        montos = re.findall(r'\b(\d{1,3}(?:\.\d{3})+)\b', texto)
        valores = [_limpiar_monto(m) for m in montos if _limpiar_monto(m)]
        if valores:
            datos['importe_total'] = max(valores)

    # ── IVA / Impuesto ────────────────────────────────────────────
    patrones_iva = [
        r'TOTAL\s+IVA[^\d]{0,10}([\d][.\d]+)',
        r'LIQUIDACI[ÓO]N\s+IVA[^%\d]{0,20}10%[^\d]{0,10}([\d][.\d]+)',
        r'IVA\s+10%[^\d]{0,10}([\d][.\d]+)',
        r'IVA\s+5%[^\d]{0,10}([\d][.\d]+)',
    ]
    for pat in patrones_iva:
        m = re.search(pat, texto_upper)
        if m:
            val = _limpiar_monto(m.group(1))
            if val:
                datos['importe_impuesto'] = val
                break

    # ── Fecha de emisión ──────────────────────────────────────────
    # DD/MM/YYYY
    fecha_m = re.search(r'\b(\d{2})[/](\d{2})[/](\d{4})\b', texto)
    if fecha_m:
        try:
            datos['fecha_emision'] = _dt(
                int(fecha_m.group(3)), int(fecha_m.group(2)), int(fecha_m.group(1))
            ).date()
        except ValueError:
            pass

    if not datos['fecha_emision']:
        # DD/MM/YY
        fecha_m2 = re.search(r'\b(\d{2})[/](\d{2})[/](\d{2})\b', texto)
        if fecha_m2:
            try:
                datos['fecha_emision'] = _dt(
                    2000 + int(fecha_m2.group(3)), int(fecha_m2.group(2)), int(fecha_m2.group(1))
                ).date()
            except ValueError:
                pass

    if not datos['fecha_emision']:
        meses = {'ENERO':'01','FEBRERO':'02','MARZO':'03','ABRIL':'04','MAYO':'05',
                 'JUNIO':'06','JULIO':'07','AGOSTO':'08','SEPTIEMBRE':'09',
                 'OCTUBRE':'10','NOVIEMBRE':'11','DICIEMBRE':'12'}
        fm = re.search(
            r'(\d{1,2})\s+(?:DE\s+)?(ENERO|FEBRERO|MARZO|ABRIL|MAYO|JUNIO|JULIO|'
            r'AGOSTO|SEPTIEMBRE|OCTUBRE|NOVIEMBRE|DICIEMBRE)\s+(?:DE\s+)?(\d{4})',
            texto_upper
        )
        if fm:
            try:
                datos['fecha_emision'] = _dt(
                    int(fm.group(3)), int(meses[fm.group(2)]), int(fm.group(1))
                ).date()
            except ValueError:
                pass

    return datos


# ══════════════════════════════════════════════════════
# PREPROCESAMIENTO DE IMAGEN
# ══════════════════════════════════════════════════════

def preprocesar_imagen(imagen):
    """Escala de grises + contraste + nitidez para mejorar OCR."""
    from PIL import Image, ImageEnhance, ImageFilter
    import io

    img = Image.open(imagen)
    if img.mode not in ('L', 'RGB'):
        img = img.convert('RGB')
    img = img.convert('L')

    w, h = img.size
    if w < 1200:
        factor = 1200 / w
        img = img.resize((int(w * factor), int(h * factor)), Image.LANCZOS)

    img = ImageEnhance.Contrast(img).enhance(2.0)
    img = ImageEnhance.Sharpness(img).enhance(2.0)
    img = img.filter(ImageFilter.SHARPEN)

    buf = io.BytesIO()
    img.save(buf, format='PNG')
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════
# OCR TESSERACT (FALLBACK)
# ══════════════════════════════════════════════════════

def ocr_con_tesseract(imagen):
    import pytesseract
    from PIL import Image as PILImage
    import shutil, os

    cmd = settings.TESSERACT_CMD
    if not os.path.exists(cmd):
        cmd = shutil.which('tesseract') or cmd
    pytesseract.pytesseract.tesseract_cmd = cmd

    img = PILImage.open(imagen)
    config = r'--oem 3 --psm 6'
    try:
        return pytesseract.image_to_string(img, lang='spa', config=config)
    except pytesseract.TesseractError:
        return pytesseract.image_to_string(img, config=config)


# ══════════════════════════════════════════════════════
# EXTRACCIÓN CON IA (CLAUDE)
# ══════════════════════════════════════════════════════

_PROMPT_OCR = """Sos un experto en facturas paraguayas. Del texto OCR extraé los campos indicados.

TIPOS DE DOCUMENTOS:
- Factura física: Timbrado DNIT, RUC emisor, número NNN-NNN-NNNNNNN
- KuDE (electrónica): misma estructura pero formato digital

REGLAS:
- ruc: del EMISOR (quien vende), no del comprador. Formato NNNNNNN-N (ej: "4934368-8")
- timbrado: número de resolución DNIT, 7-10 dígitos (ej: "18733343")
- numero_factura: formato NNN-NNN-NNNNNNN (ej: "001-001-0000508")
- fecha_emision: convertir a YYYY-MM-DD
- importe_total: número entero sin separadores (ej: 70000, 1901004, 220000)
- importe_impuesto: total IVA, número entero. null si no figura
- Si un campo no aparece con certeza → null. No inventes.

TEXTO OCR:
{texto}

Respondé SOLO con JSON válido, sin texto adicional:
{{
  "nombre_proveedor": null,
  "ruc": null,
  "timbrado": null,
  "numero_factura": null,
  "fecha_emision": null,
  "importe_total": null,
  "importe_impuesto": null
}}"""


def extraer_con_ia(texto_ocr):
    """Segunda capa: Claude interpreta el texto OCR y devuelve campos estructurados."""
    import anthropic, json as _json

    api_key = getattr(settings, 'ANTHROPIC_API_KEY', '')
    if not api_key or not texto_ocr.strip():
        return {}

    try:
        client = anthropic.Anthropic(api_key=api_key)
        msg = client.messages.create(
            model='claude-haiku-4-5-20251001',
            max_tokens=512,
            messages=[{'role': 'user', 'content': _PROMPT_OCR.format(texto=texto_ocr[:3000])}],
        )
        respuesta = msg.content[0].text.strip()

        # Extraer JSON de la respuesta
        m = re.search(r'\{[^{}]*\}', respuesta, re.DOTALL)
        if not m:
            return {}
        datos_ia = _json.loads(m.group())

        # Convertir importe_* que vengan como strings
        for campo in ('importe_total', 'importe_impuesto'):
            v = datos_ia.get(campo)
            if isinstance(v, str):
                datos_ia[campo] = _limpiar_monto(v)

        # Convertir fecha a objeto date
        if datos_ia.get('fecha_emision'):
            from datetime import datetime as _dt
            try:
                datos_ia['fecha_emision'] = _dt.strptime(
                    datos_ia['fecha_emision'], '%Y-%m-%d'
                ).date()
            except (ValueError, TypeError):
                datos_ia['fecha_emision'] = None

        return datos_ia

    except Exception:
        return {}


# ══════════════════════════════════════════════════════
# FACTURAS - ENDPOINTS API
# ══════════════════════════════════════════════════════

class FacturaUploadView(APIView):
    """
    Pipeline OCR:
    1. Preprocesar imagen
    2. OCR.space API  (falla → Tesseract con imagen preprocesada)
    3. Extracción regex  +  Claude IA (prioridad sobre regex)
    """
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        imagen = request.FILES.get('imagen')
        if not imagen:
            return Response(
                {'error': 'No se recibió ninguna imagen.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # ── 1. Preprocesar ────────────────────────────────────────
        imagen.seek(0)
        imagen_proc = preprocesar_imagen(imagen)

        # ── 2. OCR ───────────────────────────────────────────────
        texto, fuente, aviso = '', 'api', ''
        try:
            imagen_proc.seek(0)
            texto = ocr_con_api(('factura.png', imagen_proc, 'image/png'))
        except Exception:
            try:
                imagen_proc.seek(0)
                texto = ocr_con_tesseract(imagen_proc)
                fuente = 'tesseract'
            except Exception:
                fuente = 'ninguno'
                aviso = 'OCR no disponible. Podés completar los datos manualmente.'

        if not texto.strip() and fuente != 'ninguno':
            aviso = 'OCR no pudo leer texto de la imagen. Revisá que la foto sea clara y completá los datos manualmente.'

        # ── 3. Extracción ─────────────────────────────────────────
        datos = extraer_datos_ocr(texto)

        datos_ia = extraer_con_ia(texto)
        for campo, valor in datos_ia.items():
            if valor is not None and valor != '' and campo in datos:
                datos[campo] = valor   # IA tiene prioridad sobre regex

        resp = {'texto_ocr': texto, 'datos_extraidos': datos, 'fuente_ocr': fuente}
        if aviso:
            resp['aviso'] = aviso
        return Response(resp, status=status.HTTP_200_OK)


class FacturaConfirmarView(APIView):
    """
    Guarda la factura en la DB.
    - Si viene usuario en el body, lo usa (admin puede asignar a otro usuario).
    - Si no viene, asigna el usuario logueado.
    - Si carga_manual=true, no requiere imagen.
    """
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        data = request.data.copy()

        # Asignar usuario: usa el del body si existe, sino el logueado
        if not data.get('usuario') and request.user.is_authenticated:
            data['usuario'] = request.user.id

        # Marcar como carga manual si no hay imagen
        if not request.FILES.get('imagen'):
            data['carga_manual'] = True

        serializer = FacturaSerializer(data=data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class FacturaListView(APIView):
    """Lista facturas del usuario logueado."""

    def get(self, request):
        facturas = Factura.objects.filter(
            usuario=request.user
        ).select_related('usuario')
        serializer = FacturaSerializer(facturas, many=True)
        return Response(serializer.data)


class FacturaDetailView(APIView):
    """Obtiene o edita una factura específica."""

    def _get_factura(self, pk, user):
        try:
            if _puede_ver_todo(user):
                return Factura.objects.get(pk=pk)
            return Factura.objects.get(pk=pk, usuario=user)
        except Factura.DoesNotExist:
            return None

    def get(self, request, pk):
        factura = self._get_factura(pk, request.user)
        if not factura:
            return Response({'error': 'Factura no encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        return Response(FacturaSerializer(factura).data)

    def patch(self, request, pk):
        factura = self._get_factura(pk, request.user)
        if not factura:
            return Response({'error': 'Factura no encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        serializer = FacturaSerializer(factura, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk):
        factura = self._get_factura(pk, request.user)
        if not factura:
            return Response({'error': 'Factura no encontrada.'}, status=status.HTTP_404_NOT_FOUND)
        factura.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


# ══════════════════════════════════════════════════════
# IMPORTACIÓN CSV DE FACTURAS
# ══════════════════════════════════════════════════════

# Encabezados esperados en el CSV (insensible a mayúsculas/espacios)
_CSV_COLUMNAS = ['tipo', 'timbrado', 'ruc', 'nombre_proveedor', 'numero_factura', 'total_factura', 'total_iva']

_TIPO_MAP = {
    'ELECTRONICA': 'electronica',
    'NO ELECTRONICA': 'no_electronica',
    'NO_ELECTRONICA': 'no_electronica',
}


def _normalizar_monto(valor):
    """Convierte '1.500.000' o '1500000' o '1,500,000' a Decimal."""
    valor = valor.strip().replace(' ', '')
    # Si tiene punto como separador de miles (ej: 1.500.000) → quitar puntos
    # Si tiene coma como separador decimal (ej: 1500,50) → reemplazar por punto
    # Heurística: si hay más de un punto, son separadores de miles
    if valor.count('.') > 1:
        valor = valor.replace('.', '')
    elif valor.count('.') == 1 and valor.count(',') == 0:
        # puede ser separador decimal o de miles — si hay 3 dígitos después del punto, es miles
        partes = valor.split('.')
        if len(partes[1]) == 3:
            valor = valor.replace('.', '')
    valor = valor.replace(',', '.')
    return valor


class FacturaImportCSVView(APIView):
    """
    Importa facturas desde un archivo CSV separado por punto y coma (;).

    Columnas requeridas (en cualquier orden, insensibles a mayúsculas):
        TIPO ; fecha_emision ; timbrado ; RUC ; nombre_proveedor ; numero_factura ; total_factura ; total_IVA

    TIPO aceptados: ELECTRONICA, NO ELECTRONICA, NO_ELECTRONICA
    fecha_emision: formato DD/MM/YYYY
    """
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        if not request.user.is_authenticated:
            return Response({'error': 'Debe iniciar sesión.'}, status=status.HTTP_401_UNAUTHORIZED)

        archivo = request.FILES.get('archivo')
        if not archivo:
            return Response({'error': 'No se recibió ningún archivo.'}, status=status.HTTP_400_BAD_REQUEST)

        if not archivo.name.lower().endswith('.csv'):
            return Response({'error': 'El archivo debe tener extensión .csv.'}, status=status.HTTP_400_BAD_REQUEST)

        # Leer el contenido detectando la codificación
        contenido = archivo.read()
        for encoding in ('utf-8-sig', 'utf-8', 'latin-1'):
            try:
                texto = contenido.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            return Response({'error': 'No se pudo decodificar el archivo. Use UTF-8 o Latin-1.'}, status=status.HTTP_400_BAD_REQUEST)

        reader = csv.DictReader(io.StringIO(texto), delimiter=';')

        # Normalizar nombres de columna
        if reader.fieldnames is None:
            return Response({'error': 'El archivo CSV está vacío o no tiene encabezados.'}, status=status.HTTP_400_BAD_REQUEST)

        col_map = {col.strip().lower(): col for col in reader.fieldnames}
        columnas_requeridas = {
            'tipo': None, 'fecha_emision': None, 'timbrado': None, 'ruc': None,
            'nombre_proveedor': None, 'numero_factura': None,
            'total_factura': None, 'total_iva': None,
        }
        alias = {
            'fecha de emision': 'fecha_emision',
            'fecha de emisión': 'fecha_emision',
            'fecha': 'fecha_emision',
            'nombre del proveedor': 'nombre_proveedor',
            'numero factura': 'numero_factura',
            'número factura': 'numero_factura',
            'total factura': 'total_factura',
            'total iva': 'total_iva',
            'iva': 'total_iva',
        }
        for col_raw, col_original in col_map.items():
            key = alias.get(col_raw, col_raw)
            if key in columnas_requeridas:
                columnas_requeridas[key] = col_original

        faltantes = [k for k, v in columnas_requeridas.items() if v is None]
        if faltantes:
            return Response(
                {'error': f'Faltan las columnas: {", ".join(faltantes)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        importadas = []
        errores = []

        for num_fila, fila in enumerate(reader, start=2):  # start=2 porque fila 1 es encabezado
            def campo(key):
                return (fila.get(columnas_requeridas[key]) or '').strip()

            tipo_raw = campo('tipo').upper()
            tipo = _TIPO_MAP.get(tipo_raw)
            if not tipo:
                errores.append({
                    'fila': num_fila,
                    'error': f'TIPO inválido: "{campo("tipo")}". Use ELECTRONICA o NO ELECTRONICA.'
                })
                continue

            timbrado = campo('timbrado')
            ruc = campo('ruc')
            nombre_proveedor = campo('nombre_proveedor')
            numero_factura = campo('numero_factura')
            total_factura_str = campo('total_factura')
            total_iva_str = campo('total_iva')
            fecha_str = campo('fecha_emision')

            if not nombre_proveedor:
                errores.append({'fila': num_fila, 'error': 'nombre_proveedor está vacío.'})
                continue

            # Parsear fecha DD/MM/YYYY
            from datetime import datetime as dt
            fecha_emision = None
            if fecha_str:
                try:
                    fecha_emision = dt.strptime(fecha_str, '%d/%m/%Y').date()
                except ValueError:
                    errores.append({'fila': num_fila, 'error': f'fecha_emision inválida: "{fecha_str}". Use DD/MM/YYYY.'})
                    continue

            try:
                importe_total = _normalizar_monto(total_factura_str) if total_factura_str else None
                importe_total = float(importe_total) if importe_total else None
            except (ValueError, AttributeError):
                errores.append({'fila': num_fila, 'error': f'total_factura inválido: "{total_factura_str}".'})
                continue

            try:
                importe_impuesto = _normalizar_monto(total_iva_str) if total_iva_str else None
                importe_impuesto = float(importe_impuesto) if importe_impuesto else None
            except (ValueError, AttributeError):
                errores.append({'fila': num_fila, 'error': f'total_iva inválido: "{total_iva_str}".'})
                continue

            if ruc and timbrado and numero_factura:
                if Factura.objects.filter(
                    usuario=request.user,
                    ruc=ruc,
                    timbrado=timbrado,
                    numero_factura=numero_factura,
                ).exists():
                    errores.append({
                        'fila': num_fila,
                        'error': f'Duplicado: ya existe una factura con RUC {ruc}, timbrado {timbrado} y número {numero_factura}.',
                    })
                    continue

            factura = Factura(
                usuario=request.user,
                tipo=tipo,
                fecha_emision=fecha_emision,
                timbrado=timbrado,
                ruc=ruc,
                nombre_proveedor=nombre_proveedor,
                numero_factura=numero_factura,
                importe_total=importe_total,
                importe_impuesto=importe_impuesto,
                carga_manual=True,
            )
            factura.save()
            importadas.append({
                'fila': num_fila,
                'id': factura.pk,
                'nombre_proveedor': nombre_proveedor,
                'importe_total': importe_total,
            })

        return Response({
            'importadas': len(importadas),
            'errores': len(errores),
            'detalle_errores': errores,
            'detalle_importadas': importadas,
        }, status=status.HTTP_200_OK)


# ══════════════════════════════════════════════════════
# REPORTE DE GASTOS — lista detallada con filtros
# ══════════════════════════════════════════════════════

class FacturaReporteGastosView(APIView):
    """
    Lista paginada de facturas con filtros combinables.

    Parámetros GET:
      - desde           YYYY-MM-DD  (fecha_emision >=)
      - hasta           YYYY-MM-DD  (fecha_emision <=)
      - ruc             texto parcial
      - proveedor       texto parcial
      - timbrado        texto parcial
      - numero_factura  texto parcial
      - tipo            electronica | no_electronica
      - usuario_id      int  (solo superuser)
      - page            int  (default 1)
      - page_size       int  (default 50, max 200)
    """

    def get(self, request):
        if not request.user.is_authenticated:
            return Response({'error': 'No autenticado.'}, status=status.HTTP_401_UNAUTHORIZED)

        qs = Factura.objects.select_related('usuario').order_by('-fecha_emision', '-creado')

        # Permisos: superuser ve todos, usuario normal solo los propios
        if _puede_ver_todo(request.user):
            usuario_id = request.GET.get('usuario_id')
            if usuario_id:
                qs = qs.filter(usuario_id=usuario_id)
        else:
            qs = qs.filter(usuario=request.user)

        # Filtros
        desde = request.GET.get('desde')
        hasta = request.GET.get('hasta')
        ruc = request.GET.get('ruc', '').strip()
        proveedor = request.GET.get('proveedor', '').strip()
        timbrado = request.GET.get('timbrado', '').strip()
        numero_factura = request.GET.get('numero_factura', '').strip()
        tipo_filtro = request.GET.get('tipo', '').strip()

        if desde:
            qs = qs.filter(fecha_emision__gte=desde)
        if hasta:
            qs = qs.filter(fecha_emision__lte=hasta)
        if ruc:
            qs = qs.filter(ruc__icontains=ruc)
        if proveedor:
            qs = qs.filter(nombre_proveedor__icontains=proveedor)
        if timbrado:
            qs = qs.filter(timbrado__icontains=timbrado)
        if numero_factura:
            qs = qs.filter(numero_factura__icontains=numero_factura)
        if tipo_filtro:
            qs = qs.filter(tipo=tipo_filtro)

        total = qs.count()
        total_importe = qs.aggregate(s=Sum('importe_total'))['s'] or 0

        # Paginación
        try:
            page = max(1, int(request.GET.get('page', 1)))
            page_size = min(200, max(1, int(request.GET.get('page_size', 50))))
        except (ValueError, TypeError):
            page, page_size = 1, 50

        offset = (page - 1) * page_size
        facturas = qs[offset:offset + page_size]

        def ruc_sin_dv(ruc_str):
            """Devuelve solo la parte numérica antes del guion."""
            return ruc_str.split('-')[0].strip() if ruc_str else ''

        items = []
        for f in facturas:
            items.append({
                'id': f.pk,
                'usuario': f.usuario.get_full_name() or f.usuario.username if f.usuario else '—',
                'tipo': f.get_tipo_display(),
                'fecha_emision': str(f.fecha_emision) if f.fecha_emision else None,
                'timbrado': f.timbrado,
                'numero_factura': f.numero_factura,
                'ruc': ruc_sin_dv(f.ruc),
                'nombre_proveedor': f.nombre_proveedor,
                'importe_total': float(f.importe_total) if f.importe_total is not None else None,
                'importe_impuesto': float(f.importe_impuesto) if f.importe_impuesto is not None else None,
                'estado': f.get_estado_display(),
            })

        return Response({
            'total': total,
            'total_importe': float(total_importe),
            'page': page,
            'page_size': page_size,
            'pages': -(-total // page_size),   # ceil division
            'items': items,
        })


# ══════════════════════════════════════════════════════
# EXPORTACIÓN EXCEL DE GASTOS
# ══════════════════════════════════════════════════════

def _ruc_sin_dv(ruc_str):
    return ruc_str.split('-')[0].strip() if ruc_str else ''


class FacturaExportExcelView(APIView):
    """
    Descarga un archivo Excel con las facturas filtradas.

    Parámetros GET:
      - desde        YYYY-MM-DD
      - hasta        YYYY-MM-DD
      - usuario_id   int (solo superuser)
    """

    def get(self, request):
        if not request.user.is_authenticated:
            return Response({'error': 'No autenticado.'}, status=status.HTTP_401_UNAUTHORIZED)

        qs = Factura.objects.select_related('usuario').order_by('fecha_emision', 'nombre_proveedor')

        if _puede_ver_todo(request.user):
            usuario_id = request.GET.get('usuario_id')
            if usuario_id:
                qs = qs.filter(usuario_id=usuario_id)
        else:
            qs = qs.filter(usuario=request.user)

        desde = request.GET.get('desde')
        hasta = request.GET.get('hasta')
        if desde:
            qs = qs.filter(fecha_emision__gte=desde)
        if hasta:
            qs = qs.filter(fecha_emision__lte=hasta)

        # ── Construir Excel ────────────────────────────────────────────────────
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Gastos'

        # Estilos
        hdr_fill   = PatternFill('solid', fgColor='1A73E8')
        hdr_font   = Font(bold=True, color='FFFFFF', size=11)
        hdr_align  = Alignment(horizontal='center', vertical='center')
        total_fill = PatternFill('solid', fgColor='E8F0FE')
        total_font = Font(bold=True, size=11)
        thin = Side(style='thin', color='CCCCCC')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Encabezados
        columnas = ['Usuario', 'Tipo', 'Fecha', 'Timbrado', 'N° Factura', 'RUC', 'Proveedor', 'Importe (Gs.)']
        for col_idx, titulo in enumerate(columnas, 1):
            cell = ws.cell(row=1, column=col_idx, value=titulo)
            cell.fill  = hdr_fill
            cell.font  = hdr_font
            cell.alignment = hdr_align
            cell.border = border

        ws.row_dimensions[1].height = 22

        # Datos
        total_importe = 0
        for row_idx, f in enumerate(qs, start=2):
            usuario_str = ''
            if f.usuario:
                usuario_str = f.usuario.get_full_name() or f.usuario.username
            importe = float(f.importe_total) if f.importe_total is not None else 0
            total_importe += importe

            fila = [
                usuario_str,
                f.get_tipo_display(),
                f.fecha_emision.strftime('%d/%m/%Y') if f.fecha_emision else '',
                f.timbrado,
                f.numero_factura,
                _ruc_sin_dv(f.ruc),
                f.nombre_proveedor,
                importe,
            ]
            for col_idx, valor in enumerate(fila, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=valor)
                cell.border = border
                cell.alignment = Alignment(vertical='center')
                if col_idx == 8:  # Importe
                    cell.number_format = '#,##0'

        # Fila total
        total_row = qs.count() + 2
        ws.cell(row=total_row, column=7, value='TOTAL').font = total_font
        ws.cell(row=total_row, column=7).fill = total_fill
        ws.cell(row=total_row, column=7).alignment = Alignment(horizontal='right')
        total_cell = ws.cell(row=total_row, column=8, value=total_importe)
        total_cell.font = total_font
        total_cell.fill = total_fill
        total_cell.number_format = '#,##0'

        # Anchos de columna
        anchos = [22, 18, 14, 14, 20, 14, 32, 18]
        for i, ancho in enumerate(anchos, 1):
            ws.column_dimensions[get_column_letter(i)].width = ancho

        ws.freeze_panes = 'A2'

        # Nombre de archivo
        from datetime import date
        sufijo = ''
        if desde or hasta:
            sufijo = f"_{desde or ''}_{hasta or ''}"
        filename = f"gastos{sufijo}_{date.today().strftime('%Y%m%d')}.xlsx"

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


# ══════════════════════════════════════════════════════
# EXPORTACIÓN CSV DE GASTOS
# ══════════════════════════════════════════════════════

class FacturaExportCSVView(APIView):
    def get(self, request):
        if not request.user.is_authenticated:
            return Response({'error': 'No autenticado.'}, status=status.HTTP_401_UNAUTHORIZED)

        qs = Factura.objects.select_related('usuario').order_by('fecha_emision', 'nombre_proveedor')

        if _puede_ver_todo(request.user):
            usuario_id = request.GET.get('usuario_id')
            if usuario_id:
                qs = qs.filter(usuario_id=usuario_id)
        else:
            qs = qs.filter(usuario=request.user)

        desde = request.GET.get('desde')
        hasta = request.GET.get('hasta')
        if desde:
            qs = qs.filter(fecha_emision__gte=desde)
        if hasta:
            qs = qs.filter(fecha_emision__lte=hasta)

        from datetime import date
        sufijo = f"_{desde or ''}_{hasta or ''}" if (desde or hasta) else ''
        filename = f"gastos{sufijo}_{date.today().strftime('%Y%m%d')}.csv"

        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        writer = csv.writer(response, delimiter=';')
        writer.writerow(['RUC', 'Proveedor', 'Fecha', 'Timbrado', 'N° Factura', 'Monto', 'Impuesto', 'Tipo'])
        for f in qs:
            writer.writerow([
                _ruc_sin_dv(f.ruc),
                f.nombre_proveedor,
                f.fecha_emision.strftime('%d/%m/%Y') if f.fecha_emision else '',
                f.timbrado,
                f.numero_factura,
                float(f.importe_total) if f.importe_total is not None else '',
                float(f.importe_impuesto) if f.importe_impuesto is not None else '',
                f.get_tipo_display(),
            ])
        return response


# ══════════════════════════════════════════════════════
# EXPORTACIÓN CSV DE INGRESOS
# ══════════════════════════════════════════════════════

class IngresoExportCSVView(APIView):
    def get(self, request):
        if not request.user.is_authenticated:
            return Response({'error': 'No autenticado.'}, status=status.HTTP_401_UNAUTHORIZED)

        qs = Ingreso.objects.select_related('usuario').order_by('fecha', 'descripcion')

        if _puede_ver_todo(request.user):
            usuario_id = request.GET.get('usuario_id')
            if usuario_id:
                qs = qs.filter(usuario_id=usuario_id)
        else:
            qs = qs.filter(usuario=request.user)

        desde = request.GET.get('desde')
        hasta = request.GET.get('hasta')
        if desde:
            qs = qs.filter(fecha__gte=desde)
        if hasta:
            qs = qs.filter(fecha__lte=hasta)

        from datetime import date
        sufijo = f"_{desde or ''}_{hasta or ''}" if (desde or hasta) else ''
        filename = f"ingresos{sufijo}_{date.today().strftime('%Y%m%d')}.csv"

        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        writer = csv.writer(response, delimiter=';')
        writer.writerow(['Descripción', 'Monto', 'Categoría', 'Fecha', 'Notas', 'Usuario'])
        for i in qs:
            writer.writerow([
                i.descripcion,
                float(i.monto),
                i.get_categoria_display(),
                i.fecha.strftime('%d/%m/%Y') if i.fecha else '',
                i.notas,
                i.usuario.get_full_name() or i.usuario.username if i.usuario else '',
            ])
        return response


# ══════════════════════════════════════════════════════
# INGRESOS - ENDPOINTS API  ✅ NUEVO
# ══════════════════════════════════════════════════════

class IngresoListCreateView(APIView):
    """
    Lista y crea ingresos del usuario logueado.

    GET sin ?paginar: devuelve lista completa (IngresoSerializer) — compatible con home.html.
    GET con ?paginar=1: devuelve formato paginado con filtros.

    Filtros soportados (con ?paginar=1):
      - desde, hasta   YYYY-MM-DD
      - categoria      sueldo|freelance|venta|alquiler|transferencia|otro
      - descripcion    texto parcial
      - usuario_id     int (solo superuser)
      - page, page_size
    """

    def get(self, request):
        if not request.user.is_authenticated:
            return Response({'error': 'No autenticado.'}, status=status.HTTP_401_UNAUTHORIZED)

        qs = Ingreso.objects.select_related('usuario').order_by('-fecha', '-creado')

        if _puede_ver_todo(request.user):
            usuario_id = request.GET.get('usuario_id')
            if usuario_id:
                qs = qs.filter(usuario_id=usuario_id)
        else:
            qs = qs.filter(usuario=request.user)

        desde = request.GET.get('desde', '').strip()
        hasta = request.GET.get('hasta', '').strip()
        categoria = request.GET.get('categoria', '').strip()
        descripcion = request.GET.get('descripcion', '').strip()

        if desde:
            qs = qs.filter(fecha__gte=desde)
        if hasta:
            qs = qs.filter(fecha__lte=hasta)
        if categoria:
            qs = qs.filter(categoria=categoria)
        if descripcion:
            qs = qs.filter(descripcion__icontains=descripcion)

        if request.GET.get('paginar'):
            total = qs.count()
            total_monto = qs.aggregate(s=Sum('monto'))['s'] or 0
            try:
                page = max(1, int(request.GET.get('page', 1)))
                page_size = min(200, max(1, int(request.GET.get('page_size', 50))))
            except (ValueError, TypeError):
                page, page_size = 1, 50

            offset = (page - 1) * page_size
            items = []
            for i in qs[offset:offset + page_size]:
                items.append({
                    'id': i.pk,
                    'usuario': i.usuario.get_full_name() or i.usuario.username if i.usuario else '—',
                    'descripcion': i.descripcion,
                    'monto': float(i.monto),
                    'categoria': i.categoria,
                    'categoria_display': i.get_categoria_display(),
                    'fecha': str(i.fecha),
                    'notas': i.notas,
                })
            return Response({
                'total': total,
                'total_monto': float(total_monto),
                'page': page,
                'page_size': page_size,
                'pages': -(-total // page_size),
                'items': items,
            })

        serializer = IngresoSerializer(qs, many=True)
        return Response(serializer.data)

    def post(self, request):
        data = request.data.copy()
        data['usuario'] = request.user.id
        serializer = IngresoSerializer(data=data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class IngresoDetailView(APIView):
    """Obtiene, edita o elimina un ingreso específico."""

    def _get_ingreso(self, pk, user):
        try:
            if _puede_ver_todo(user):
                return Ingreso.objects.get(pk=pk)
            return Ingreso.objects.get(pk=pk, usuario=user)
        except Ingreso.DoesNotExist:
            return None

    def get(self, request, pk):
        ingreso = self._get_ingreso(pk, request.user)
        if not ingreso:
            return Response(
                {'error': 'Ingreso no encontrado.'},
                status=status.HTTP_404_NOT_FOUND
            )
        return Response(IngresoSerializer(ingreso).data)

    def patch(self, request, pk):
        ingreso = self._get_ingreso(pk, request.user)
        if not ingreso:
            return Response(
                {'error': 'Ingreso no encontrado.'},
                status=status.HTTP_404_NOT_FOUND
            )
        serializer = IngresoSerializer(ingreso, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, pk):
        ingreso = self._get_ingreso(pk, request.user)
        if not ingreso:
            return Response(
                {'error': 'Ingreso no encontrado.'},
                status=status.HTTP_404_NOT_FOUND
            )
        ingreso.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


# ══════════════════════════════════════════════════════
# DASHBOARD - RESUMEN FINANCIERO  ✅ NUEVO
# ══════════════════════════════════════════════════════

class DashboardView(APIView):
    """
    Retorna resumen financiero del usuario:
    - Balance mensual (ingresos - egresos del mes actual)
    - Totales anuales de ingresos y egresos
    - Últimas facturas e ingresos
    """

    def get(self, request):
        hoy = timezone.now().date()
        mes_actual = hoy.month
        anio_actual = hoy.year

        usuario = request.user

        # ── Egresos (facturas) ───────────────────────────────────────────────
        egresos_mes = Factura.objects.filter(
            usuario=usuario,
            fecha_emision__year=anio_actual,
            fecha_emision__month=mes_actual,
        ).aggregate(total=Sum('importe_total'))['total'] or 0

        egresos_anio = Factura.objects.filter(
            usuario=usuario,
            fecha_emision__year=anio_actual,
        ).aggregate(total=Sum('importe_total'))['total'] or 0

        # ── Ingresos ─────────────────────────────────────────────────────────
        ingresos_mes = Ingreso.objects.filter(
            usuario=usuario,
            fecha__year=anio_actual,
            fecha__month=mes_actual,
        ).aggregate(total=Sum('monto'))['total'] or 0

        ingresos_anio = Ingreso.objects.filter(
            usuario=usuario,
            fecha__year=anio_actual,
        ).aggregate(total=Sum('monto'))['total'] or 0

        # ── Balance ──────────────────────────────────────────────────────────
        balance_mes = ingresos_mes - egresos_mes
        balance_anio = ingresos_anio - egresos_anio

        # ── Últimos registros ─────────────────────────────────────────────────
        ultimas_facturas = Factura.objects.filter(
            usuario=usuario
        ).order_by('-creado')[:5]

        ultimos_ingresos = Ingreso.objects.filter(
            usuario=usuario
        ).order_by('-creado')[:5]

        return Response({
            'periodo': {
                'mes': mes_actual,
                'anio': anio_actual,
            },
            'mensual': {
                'ingresos': float(ingresos_mes),
                'egresos': float(egresos_mes),
                'balance': float(balance_mes),
            },
            'anual': {
                'ingresos': float(ingresos_anio),
                'egresos': float(egresos_anio),
                'balance': float(balance_anio),
            },
            'ultimas_facturas': FacturaSerializer(
                ultimas_facturas, many=True
            ).data,
            'ultimos_ingresos': IngresoSerializer(
                ultimos_ingresos, many=True
            ).data,
        })


# ══════════════════════════════════════════════════════
# USUARIOS - Endpoint para selector de usuario
# ══════════════════════════════════════════════════════

class UsuarioListView(APIView):
    """Lista usuarios disponibles. Solo devuelve todos si el usuario tiene permiso ver_todo."""

    def get(self, request):
        if not request.user.is_authenticated:
            return Response([], status=status.HTTP_200_OK)
        if _puede_ver_todo(request.user):
            usuarios = User.objects.filter(is_active=True).order_by('username')
        else:
            usuarios = User.objects.filter(pk=request.user.pk)
        serializer = UsuarioSerializer(usuarios, many=True)
        return Response(serializer.data)


# ══════════════════════════════════════════════════════
# VISTAS DJANGO (HTML/PWA) - Sin cambios estructurales
# ══════════════════════════════════════════════════════

def login_view(request):
    if request.user.is_authenticated:
        return redirect('home')
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('home')
        else:
            messages.error(request, 'Usuario o contraseña incorrectos.')
    return render(request, 'pwa/login.html')


def logout_view(request):
    logout(request)
    return redirect('login')


@login_required(login_url='login')
def home_view(request):
    facturas = Factura.objects.filter(
        usuario=request.user
    ).order_by('-creado')[:50]
    context = {'facturas': facturas}
    return render(request, 'pwa/home.html', context)

@login_required(login_url='login')
def reporte_view(request):
    """Vista HTML para la página de reportería."""
    return render(request, 'pwa/reporte.html')

@login_required(login_url='login')
def consulta_view(request):
    """Vista HTML para consulta y gestión de registros."""
    return render(request, 'pwa/consulta.html')


# ══════════════════════════════════════════════════════
# GOOGLE SHEETS — OAUTH2 + SINCRONIZACIÓN
# ══════════════════════════════════════════════════════

_SHEETS_SCOPES = ['https://www.googleapis.com/auth/spreadsheets']

def _GOOGLE_CLIENT_CONFIG():
    # 'installed' = credencial tipo Desktop/Escritorio en Google Cloud Console.
    # Para localhost no requiere registrar redirect URIs (Google lo permite automáticamente).
    # Para producción (Render) sí hay que registrar la URI.
    cred_type = getattr(settings, 'GOOGLE_CREDENTIAL_TYPE', 'installed')
    return {
        cred_type: {
            'client_id':     settings.GOOGLE_CLIENT_ID,
            'client_secret': settings.GOOGLE_CLIENT_SECRET,
            'redirect_uris': [settings.GOOGLE_REDIRECT_URI],
            'auth_uri':  'https://accounts.google.com/o/oauth2/auth',
            'token_uri': 'https://oauth2.googleapis.com/token',
        }
    }


def _build_flow():
    from google_auth_oauthlib.flow import Flow
    flow = Flow.from_client_config(
        client_config=_GOOGLE_CLIENT_CONFIG(),
        scopes=_SHEETS_SCOPES,
    )
    flow.redirect_uri = settings.GOOGLE_REDIRECT_URI
    return flow


def _get_gspread_client(config_obj):
    """Devuelve un cliente gspread autenticado, refrescando token si es necesario."""
    from google.oauth2.credentials import Credentials
    from google.auth.transport.requests import Request, AuthorizedSession
    import gspread

    creds = Credentials(
        token=config_obj.access_token,
        refresh_token=config_obj.refresh_token,
        token_uri='https://oauth2.googleapis.com/token',
        client_id=settings.GOOGLE_CLIENT_ID,
        client_secret=settings.GOOGLE_CLIENT_SECRET,
        scopes=_SHEETS_SCOPES,
    )

    if not creds.valid and creds.refresh_token:
        creds.refresh(Request())
        config_obj.access_token = creds.token
        if creds.expiry:
            from datetime import timezone as _tz
            expiry = creds.expiry
            if expiry.tzinfo is None:
                expiry = expiry.replace(tzinfo=_tz.utc)
            config_obj.token_expiry = expiry
        config_obj.save(update_fields=['access_token', 'token_expiry'])

    gc = gspread.Client(auth=creds)
    gc.session = AuthorizedSession(creds)
    return gc


def _extract_spreadsheet_id(url_or_id: str) -> str | None:
    """Extrae el ID de una URL de Google Sheets o lo devuelve si ya es un ID."""
    import re
    if not url_or_id:
        return None
    m = re.search(r'/spreadsheets/d/([a-zA-Z0-9\-_]+)', url_or_id)
    if m:
        return m.group(1)
    if re.match(r'^[a-zA-Z0-9\-_]{20,}$', url_or_id):
        return url_or_id
    return None


def _escribir_hoja_gastos(sh, ws_name, qs):
    import gspread
    headers = ['Usuario', 'Fecha', 'Timbrado', 'N° Factura', 'RUC', 'Proveedor',
               'Tipo', 'Importe (Gs.)', 'IVA (Gs.)', 'Estado', 'Notas']
    try:
        ws = sh.worksheet(ws_name)
        ws.clear()
    except gspread.exceptions.WorksheetNotFound:
        ws = sh.add_worksheet(title=ws_name, rows=max(qs.count() + 10, 50), cols=11)
    rows = [headers]
    for f in qs:
        rows.append([
            f.usuario.get_full_name() or f.usuario.username if f.usuario else '',
            f.fecha_emision.strftime('%d/%m/%Y') if f.fecha_emision else '',
            f.timbrado or '',
            f.numero_factura or '',
            f.ruc.split('-')[0].strip() if f.ruc else '',
            f.nombre_proveedor or '',
            f.get_tipo_display(),
            float(f.importe_total) if f.importe_total is not None else '',
            float(f.importe_impuesto) if f.importe_impuesto is not None else '',
            f.get_estado_display(),
            f.notas or '',
        ])
    ws.update(rows, 'A1')


def _escribir_hoja_ingresos(sh, ws_name, qs):
    import gspread
    headers = ['Usuario', 'Fecha', 'Descripción', 'Categoría', 'Monto (Gs.)', 'Notas']
    try:
        ws = sh.worksheet(ws_name)
        ws.clear()
    except gspread.exceptions.WorksheetNotFound:
        ws = sh.add_worksheet(title=ws_name, rows=max(qs.count() + 10, 50), cols=6)
    rows = [headers]
    for i in qs:
        rows.append([
            i.usuario.get_full_name() or i.usuario.username if i.usuario else '',
            i.fecha.strftime('%d/%m/%Y') if i.fecha else '',
            i.descripcion,
            i.get_categoria_display(),
            float(i.monto),
            i.notas or '',
        ])
    ws.update(rows, 'A1')


def _sincronizar_datos(user, config_obj):
    """
    Sincroniza a Google Sheets:
    - Hojas consolidadas: 'Gastos YYYY', 'Ingresos YYYY'
    - Hojas por usuario:  'Gastos YYYY username', 'Ingresos YYYY username'
    Para usuarios con ver_todo: consolida todos los usuarios en las hojas generales.
    """
    gc = _get_gspread_client(config_obj)
    sh = gc.open_by_key(config_obj.spreadsheet_id)

    ver_todo = _puede_ver_todo(user)
    usuarios = list(
        User.objects.filter(
            Q(facturas__isnull=False) | Q(ingresos__isnull=False)
        ).distinct().order_by('username')
    ) if ver_todo else [user]

    hojas_escritas = []

    factura_qs_base = Factura.objects.filter(usuario__in=usuarios, fecha_emision__isnull=False).select_related('usuario')
    ingreso_qs_base = Ingreso.objects.filter(usuario__in=usuarios).select_related('usuario')

    # ── Hojas consolidadas por año ────────────────────────────────
    for year in sorted(set(factura_qs_base.values_list('fecha_emision__year', flat=True)), reverse=True):
        qs = factura_qs_base.filter(fecha_emision__year=year).order_by('fecha_emision', 'nombre_proveedor')
        ws_name = f'Gastos {year}'
        _escribir_hoja_gastos(sh, ws_name, qs)
        hojas_escritas.append(ws_name)

    for year in sorted(set(ingreso_qs_base.values_list('fecha__year', flat=True)), reverse=True):
        qs = ingreso_qs_base.filter(fecha__year=year).order_by('fecha', 'descripcion')
        ws_name = f'Ingresos {year}'
        _escribir_hoja_ingresos(sh, ws_name, qs)
        hojas_escritas.append(ws_name)

    # ── Hojas individuales por usuario ────────────────────────────
    for u in usuarios:
        uname = u.username

        for year in sorted(set(
            Factura.objects.filter(usuario=u, fecha_emision__isnull=False)
            .values_list('fecha_emision__year', flat=True)
        ), reverse=True):
            qs = Factura.objects.filter(usuario=u, fecha_emision__year=year).select_related('usuario').order_by('fecha_emision', 'nombre_proveedor')
            ws_name = f'Gastos {year} {uname}'
            _escribir_hoja_gastos(sh, ws_name, qs)
            hojas_escritas.append(ws_name)

        for year in sorted(set(
            Ingreso.objects.filter(usuario=u)
            .values_list('fecha__year', flat=True)
        ), reverse=True):
            qs = Ingreso.objects.filter(usuario=u, fecha__year=year).select_related('usuario').order_by('fecha', 'descripcion')
            ws_name = f'Ingresos {year} {uname}'
            _escribir_hoja_ingresos(sh, ws_name, qs)
            hojas_escritas.append(ws_name)

    config_obj.ultima_sincronizacion = timezone.now()
    config_obj.save(update_fields=['ultima_sincronizacion'])

    return {
        'hojas': hojas_escritas,
        'total_facturas': factura_qs_base.count(),
        'total_ingresos': ingreso_qs_base.count(),
    }


# ── Vistas OAuth ──────────────────────────────────────────────────────────────

@login_required(login_url='login')
def sheets_debug(request):
    """Solo en DEBUG: muestra la URL OAuth que se generaría."""
    if not settings.DEBUG:
        from django.http import Http404
        raise Http404
    import urllib.parse as up
    flow = _build_flow()
    url, _ = flow.authorization_url(access_type='offline', prompt='consent')
    params = dict(up.parse_qsl(up.urlparse(url).query))
    from django.http import JsonResponse
    return JsonResponse({
        'redirect_uri_que_envia_la_app': params.get('redirect_uri'),
        'client_id': params.get('client_id'),
        'GOOGLE_REDIRECT_URI_en_settings': settings.GOOGLE_REDIRECT_URI,
        'url_completa': url,
        'instruccion': 'Registrá exactamente "redirect_uri_que_envia_la_app" en Google Cloud Console',
    }, json_dumps_params={'ensure_ascii': False, 'indent': 2})


@login_required(login_url='login')
def sheets_conectar(request):
    """Inicia el flujo OAuth2 con Google."""
    if not settings.GOOGLE_CLIENT_ID:
        messages.error(request, 'Google Sheets no está configurado en este servidor.')
        return redirect('reporte')

    # Necesario para permitir redirect HTTP en desarrollo local
    if settings.DEBUG:
        import os
        os.environ['OAUTHLIB_INSECURE_TRANSPORT'] = '1'

    flow = _build_flow()
    authorization_url, state = flow.authorization_url(
        access_type='offline',
        prompt='consent',
    )
    request.session['google_oauth_state'] = state
    return redirect(authorization_url)


@login_required(login_url='login')
def sheets_callback(request):
    """Maneja el callback de Google OAuth2 y almacena los tokens."""
    if not request.user.is_authenticated:
        return redirect('login')

    error = request.GET.get('error')
    if error:
        messages.error(request, f'Acceso denegado por Google: {error}')
        return redirect('reporte')

    try:
        flow = _build_flow()
        flow.fetch_token(authorization_response=request.build_absolute_uri())
        creds = flow.credentials

        from datetime import timezone as _tz
        expiry = creds.expiry
        if expiry and expiry.tzinfo is None:
            expiry = expiry.replace(tzinfo=_tz.utc)

        from .models import GoogleSheetConfig
        cfg, _ = GoogleSheetConfig.objects.get_or_create(usuario=request.user)
        cfg.access_token  = creds.token or ''
        cfg.refresh_token = creds.refresh_token or cfg.refresh_token
        cfg.token_expiry  = expiry
        cfg.save()

        messages.success(request, '✅ Cuenta de Google conectada correctamente.')
    except Exception as e:
        messages.error(request, f'Error al conectar con Google: {e}')

    return redirect('reporte')


@login_required(login_url='login')
def sheets_desconectar(request):
    """Elimina los tokens del usuario."""
    from .models import GoogleSheetConfig
    GoogleSheetConfig.objects.filter(usuario=request.user).update(
        access_token='', refresh_token='', token_expiry=None
    )
    messages.success(request, 'Cuenta de Google desconectada.')
    return redirect('reporte')


# ── APIs ──────────────────────────────────────────────────────────────────────

class SheetsConfigView(APIView):
    """GET: estado de conexión. PATCH: actualiza spreadsheet_id."""

    def get(self, request):
        if not request.user.is_authenticated:
            return Response({'error': 'No autenticado.'}, status=status.HTTP_401_UNAUTHORIZED)

        from .models import GoogleSheetConfig
        cfg = GoogleSheetConfig.objects.filter(usuario=request.user).first()

        if not cfg:
            return Response({
                'conectado': False,
                'spreadsheet_id': '',
                'ultima_sincronizacion': None,
                'google_configurado': bool(settings.GOOGLE_CLIENT_ID),
            })

        return Response({
            'conectado': cfg.conectado,
            'spreadsheet_id': cfg.spreadsheet_id,
            'ultima_sincronizacion': cfg.ultima_sincronizacion.isoformat() if cfg.ultima_sincronizacion else None,
            'google_configurado': bool(settings.GOOGLE_CLIENT_ID),
        })

    def patch(self, request):
        if not request.user.is_authenticated:
            return Response({'error': 'No autenticado.'}, status=status.HTTP_401_UNAUTHORIZED)

        url_o_id = (request.data.get('spreadsheet_id') or '').strip()
        sheet_id = _extract_spreadsheet_id(url_o_id)
        if not sheet_id:
            return Response(
                {'error': 'URL o ID de hoja inválido.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        from .models import GoogleSheetConfig
        cfg, _ = GoogleSheetConfig.objects.get_or_create(usuario=request.user)
        cfg.spreadsheet_id = sheet_id
        cfg.save(update_fields=['spreadsheet_id'])
        return Response({'spreadsheet_id': sheet_id})


class SheetsSincronizarView(APIView):
    """Ejecuta la sincronización completa con Google Sheets."""

    def post(self, request):
        if not request.user.is_authenticated:
            return Response({'error': 'No autenticado.'}, status=status.HTTP_401_UNAUTHORIZED)

        from .models import GoogleSheetConfig
        cfg = GoogleSheetConfig.objects.filter(usuario=request.user).first()

        if not cfg or not cfg.conectado:
            return Response(
                {'error': 'Primero conectá tu cuenta de Google.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        if not cfg.spreadsheet_id:
            return Response(
                {'error': 'Ingresá la URL o ID de tu hoja de cálculo.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            resultado = _sincronizar_datos(request.user, cfg)
            return Response({
                'ok': True,
                'hojas': resultado['hojas'],
                'total_facturas': resultado['total_facturas'],
                'total_ingresos': resultado['total_ingresos'],
                'ultima_sincronizacion': cfg.ultima_sincronizacion.isoformat(),
            })
        except Exception as e:
            return Response(
                {'error': f'Error al sincronizar: {e}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class ReporteDashboardView(APIView):
    """
    Dashboard con filtro por usuario.
    Parámetros GET:
      - usuario_id: int (opcional, solo admin)
      - anio: int (opcional, default año actual)
    """

    def get(self, request):
        from django.db.models.functions import ExtractMonth
        import calendar

        anio = int(request.GET.get('anio', timezone.now().year))

        # Determinar qué usuario filtrar
        # Si es superuser puede ver cualquier usuario
        usuario_id = request.GET.get('usuario_id')
        if usuario_id and _puede_ver_todo(request.user):
            try:
                usuario_filtro = User.objects.get(pk=usuario_id)
            except User.DoesNotExist:
                usuario_filtro = request.user
        else:
            usuario_filtro = request.user

        meses_nombres = [
            'Ene','Feb','Mar','Abr','May','Jun',
            'Jul','Ago','Sep','Oct','Nov','Dic'
        ]

        # Construir tabla mensual
        mensual = []
        total_ing = 0
        total_egr = 0

        for mes in range(1, 13):
            ing = Ingreso.objects.filter(
                usuario=usuario_filtro,
                fecha__year=anio,
                fecha__month=mes,
            ).aggregate(t=Sum('monto'))['t'] or 0

            egr = Factura.objects.filter(
                usuario=usuario_filtro,
                fecha_emision__year=anio,
                fecha_emision__month=mes,
            ).aggregate(t=Sum('importe_total'))['t'] or 0

            total_ing += float(ing)
            total_egr += float(egr)

            mensual.append({
                'mes': mes,
                'nombre': meses_nombres[mes - 1],
                'ingresos': float(ing),
                'egresos': float(egr),
                'balance': float(ing) - float(egr),
            })

        # Totales generales
        mes_actual = timezone.now().month
        ing_mes = Ingreso.objects.filter(
            usuario=usuario_filtro,
            fecha__year=anio,
            fecha__month=mes_actual,
        ).aggregate(t=Sum('monto'))['t'] or 0

        egr_mes = Factura.objects.filter(
            usuario=usuario_filtro,
            fecha_emision__year=anio,
            fecha_emision__month=mes_actual,
        ).aggregate(t=Sum('importe_total'))['t'] or 0

        return Response({
            'usuario': {
                'id': usuario_filtro.id,
                'username': usuario_filtro.username,
                'nombre': usuario_filtro.get_full_name() or usuario_filtro.username,
            },
            'anio': anio,
            'mes_actual': mes_actual,
            'mensual': mensual,
            'resumen': {
                'ingresos_mes':  float(ing_mes),
                'egresos_mes':   float(egr_mes),
                'balance_mes':   float(ing_mes) - float(egr_mes),
                'ingresos_anio': total_ing,
                'egresos_anio':  total_egr,
                'balance_anio':  total_ing - total_egr,
            },
        })

